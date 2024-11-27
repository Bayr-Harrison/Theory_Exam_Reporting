import streamlit as st
import pandas as pd
import pg8000
import os
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Set up Streamlit
st.title("Vocational Assessment Theory Reporting Portal")
st.write("Enter the password to access the Reporting Portal.")

# Authentication
if "authenticated" not in st.session_state:
    st.session_state["authenticated"] = False

# Handle authentication
if not st.session_state["authenticated"]:
    password = st.text_input("Password", type="password")
    if password:
        if password == os.environ["APP_PASSWORD"]:
            st.session_state["authenticated"] = True
            st.success("Authenticated successfully!")
        else:
            st.error("Incorrect password. Please try again.")

# Main app functionality
if st.session_state["authenticated"]:
    st.write("Select a date range to query the database and download results as an Excel file.")

    # Date inputs
    start_date = st.date_input("Start Date")
    end_date = st.date_input("End Date")

    # Connect to Supabase database
    def connect_to_supabase():
        return pg8000.connect(
            database=os.environ["SUPABASE_DB_NAME"],
            user=os.environ["SUPABASE_USER"],
            password=os.environ["SUPABASE_PASSWORD"],
            host=os.environ["SUPABASE_HOST"],
            port=os.environ["SUPABASE_PORT"]
        )

    # Query the database
    def query_database(start_date, end_date):
        db_query = f"""
            SELECT student_list.name,                  
                   student_list.iatc_id,
                   exam_results.nat_id,
                   student_list.class,
                   student_list.curriculum,
                   exam_results.exam,
                   exam_results.score,
                   exam_results.result,
                   exam_results.session,
                   exam_results.date,
                   exam_results.type,
                   exam_results.attempt_index,
                   exam_results.score_index
            FROM exam_results 
            JOIN student_list ON exam_results.nat_id = student_list.nat_id
            WHERE exam_results.date >= '{start_date}' AND exam_results.date <= '{end_date}'
            ORDER BY exam_results.date ASC, exam_results.session ASC, student_list.class, student_list.iatc_id ASC
        """
        try:
            connection = connect_to_supabase()
            cursor = connection.cursor()
            cursor.execute(db_query)
            result = cursor.fetchall()
            cursor.close()
            connection.close()
            return result
        except Exception as e:
            st.error(f"Error querying database: {e}")
            return []

    # Create and download Excel file
    def create_excel(data, start_date, end_date):
        col_names = [
            'Name', 'IATC ID', 'National ID', 'Class', 'Faculty', 
            'Exam', 'Score', 'Result', 'Session', 'Date', 'Type',
            'Attempt Index', 'Score Index'
        ]
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        
        # Write the headers
        for col_num, header in enumerate(col_names, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            # Set header background color
            cell.fill = PatternFill(start_color="8DE7D3", end_color="8DE7D3", fill_type="solid")
            # Center align headers
            cell.alignment = Alignment(horizontal="center", vertical="center")
            # Bold headers
            cell.font = Font(bold=True)
        
        # Write data rows
        for row_num, row_data in enumerate(data, 2):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                # Center align all data cells
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Apply filters on the header row
        ws.auto_filter.ref = ws.dimensions

        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2
        
        # Apply borders to all data
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(col_names)):
            for cell in row:
                cell.border = thin_border
        
        # Format the Class column (Column D) as numbers with two decimal places
        for cell in ws["D"][1:]:  # Skip the header
            cell.number_format = "0.00"
        
        # Format the Date column (Column J) as 'dd-mmm-yyyy'
        for cell in ws["J"][1:]:  # Skip the header
            cell.number_format = "DD-MMM-YYYY"
        
        # Conditional formatting for Results column (Column H)
        for cell in ws["H"][1:]:  # Skip the header
            if cell.value == "PASS":
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
            elif cell.value == "FAIL":
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
        
        # Save the Excel file to a BytesIO object
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # Button to trigger query and download
    if st.button("Query and Download"):
        if start_date and end_date:
            st.write("Querying database...")
            data = query_database(start_date, end_date)
            if data:
                st.write(f"Query returned {len(data)} rows.")
                excel_file = create_excel(data, start_date, end_date)
                file_name = f"Exam Results ({start_date} to {end_date}).xlsx"
                st.download_button(
                    label="Download Excel File",
                    data=excel_file,
                    file_name=file_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                st.warning("No data found for the specified date range.")
        else:
            st.error("Please select both start and end dates.")